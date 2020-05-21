#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/5/21 16:31
# @Author  : evan_qb
# @Site    : 
# @File    : updateExcel.py
# @Software: PyCharm

import openpyxl
import os

student = {
    'name': '小红3',
    'age': 23,
    'no': 3
}

path = 'F:\\test\\excel\\'
os.chdir(path)
wb = openpyxl.load_workbook(path + 'test1.xlsx')
for sheetname in wb.sheetnames:
    sheet = wb[sheetname]
    for r in range(2, sheet.max_row + 1):
        stuName = sheet.cell(row=r, column=1).value
        stuAge = sheet.cell(row=r, column=2).value
        stuNo = sheet.cell(row=r, column=3).value
        if stuName == student['name']:
            sheet.cell(row=r, column=2).value = student['age']
            sheet.cell(row=r, column=3).value = student['no']

# 另存为test2.xlsx
wb.save('test2.xlsx')

